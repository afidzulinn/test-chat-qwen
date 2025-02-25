import os
import uuid
import logging
from typing import List, Dict, Any, Tuple
import shutil
import re

# Import document processing libraries
import PyPDF2
from docx import Document
import pandas as pd
import openpyxl

logger = logging.getLogger(__name__)

class FileProcessor:
    def __init__(self):
        self.upload_dir = "data/uploaded_files"
        os.makedirs(self.upload_dir, exist_ok=True)

    async def save_upload_file(self, file) -> str:
        """
        Save an uploaded file to disk
        
        Args:
            file: The uploaded file object
            
        Returns:
            str: Path to the saved file
        """
        try:
            # Generate a unique filename
            file_extension = os.path.splitext(file.filename)[1]
            unique_filename = f"{uuid.uuid4()}{file_extension}"
            file_path = os.path.join(self.upload_dir, unique_filename)
            
            # Save the file
            with open(file_path, "wb") as buffer:
                shutil.copyfileobj(file.file, buffer)
                
            logger.info(f"File saved: {file_path}")
            return file_path
        except Exception as e:
            logger.error(f"Error saving file: {str(e)}")
            raise

    def extract_text_from_file(self, file_path: str) -> Tuple[List[str], List[Dict[str, Any]]]:
        """
        Extract text and metadata from a file
        
        Args:
            file_path: Path to the file
            
        Returns:
            Tuple[List[str], List[Dict]]: List of text chunks and corresponding metadata
        """
        try:
            file_extension = os.path.splitext(file_path)[1].lower()
            
            if file_extension == '.pdf':
                return self._process_pdf(file_path)
            elif file_extension == '.docx':
                return self._process_docx(file_path)
            elif file_extension in ['.xlsx', '.xls']:
                return self._process_excel(file_path)
            else:
                raise ValueError(f"Unsupported file type: {file_extension}")
        except Exception as e:
            logger.error(f"Error extracting text from file: {str(e)}")
            raise

    def _process_pdf(self, file_path: str) -> Tuple[List[str], List[Dict[str, Any]]]:
        """Process PDF file and extract text with metadata"""
        texts = []
        metadatas = []
        
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                
                # Extract document info
                info = pdf_reader.metadata
                title = info.title if info and info.title else os.path.basename(file_path)
                
                # Process each page
                for i, page in enumerate(pdf_reader.pages):
                    page_text = page.extract_text()
                    
                    # Split into chunks (~ 1000 chars)
                    chunks = self._split_into_chunks(page_text, 1000)
                    
                    for j, chunk in enumerate(chunks):
                        texts.append(chunk)
                        metadatas.append({
                            "source": os.path.basename(file_path),
                            "type": "pdf",
                            "title": title,
                            "page": i + 1,
                            "chunk": j + 1
                        })
            
            return texts, metadatas
        except Exception as e:
            logger.error(f"Error processing PDF: {str(e)}")
            raise

    def _process_docx(self, file_path: str) -> Tuple[List[str], List[Dict[str, Any]]]:
        """Process DOCX file and extract text with metadata"""
        texts = []
        metadatas = []
        
        try:
            doc = Document(file_path)
            title = os.path.basename(file_path)
            
            # Extract all paragraphs
            full_text = " ".join([para.text for para in doc.paragraphs if para.text.strip()])
            
            # Split into chunks
            chunks = self._split_into_chunks(full_text, 1000)
            
            for i, chunk in enumerate(chunks):
                texts.append(chunk)
                metadatas.append({
                    "source": os.path.basename(file_path),
                    "type": "docx",
                    "title": title,
                    "chunk": i + 1
                })
            
            return texts, metadatas
        except Exception as e:
            logger.error(f"Error processing DOCX: {str(e)}")
            raise

    def _process_excel(self, file_path: str) -> Tuple[List[str], List[Dict[str, Any]]]:
        """Process Excel file and extract text with metadata"""
        texts = []
        metadatas = []
        
        try:
            # Load the workbook
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # Process each sheet
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Convert to DataFrame for easier processing
                data = []
                for row in sheet.iter_rows(values_only=True):
                    data.append(row)
                
                if not data:
                    continue
                    
                df = pd.DataFrame(data[1:], columns=data[0] if data else None)
                
                # Convert DataFrame to text representation
                text_representation = f"Sheet: {sheet_name}\n"
                text_representation += df.to_string(index=False)
                
                # Split into chunks
                chunks = self._split_into_chunks(text_representation, 1000)
                
                for i, chunk in enumerate(chunks):
                    texts.append(chunk)
                    metadatas.append({
                        "source": os.path.basename(file_path),
                        "type": "excel",
                        "title": os.path.basename(file_path),
                        "sheet": sheet_name,
                        "chunk": i + 1
                    })
            
            return texts, metadatas
        except Exception as e:
            logger.error(f"Error processing Excel: {str(e)}")
            raise

    def _split_into_chunks(self, text: str, chunk_size: int) -> List[str]:
        """Split text into chunks of approximately chunk_size characters"""
        chunks = []
        current_chunk = ""
        
        # Split by sentences to avoid cutting in the middle of a sentence
        sentences = re.split(r'(?<=[.!?])\s+', text)
        
        for sentence in sentences:
            if len(current_chunk) + len(sentence) <= chunk_size:
                current_chunk += sentence + " "
            else:
                if current_chunk:
                    chunks.append(current_chunk.strip())
                current_chunk = sentence + " "
                
        if current_chunk:
            chunks.append(current_chunk.strip())
            
        return chunks

# Create a singleton instance
file_processor = FileProcessor()